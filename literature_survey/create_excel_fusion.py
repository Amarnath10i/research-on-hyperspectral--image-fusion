import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "IEEE_2026_HSI_Fusion_Only"

papers = [
    {
        "id": 1,
        "title": "Two-Step Pansharpening: A High-Frequency-Guided Spatial-Spectral Enhancement Network Based on Mixture of Experts",
        "authors": "Li, Zhuomao; Li, Ying; Li, Zhihao; Liu, Yikun; Yang, Gongping",
        "journal": "IEEE Transactions on Geoscience and Remote Sensing (TGRS)",
        "year": 2026,
        "volume": 64,
        "pages": "1-19",
        "doi": "10.1109/tgrs.2025.3648716",
        "impact_factor": 9.4,
        "open_access": "No",
        "code": "Yes (github.com/lzm-01/TS-Pan)",
        "category": "Pansharpening (Fusion)"
    },
    {
        "id": 2,
        "title": "Frequency-Driven State-Space Model for Remote Sensing Pansharpening",
        "authors": "Zhang, Xiaochen; Li, Jiangyun; Wang, Hao; Zhuang, Peixian",
        "journal": "IEEE Transactions on Geoscience and Remote Sensing (TGRS)",
        "year": 2026,
        "volume": 64,
        "pages": "5911314",
        "doi": "10.1109/TGRS.2026.3699514",
        "impact_factor": 9.4,
        "open_access": "No",
        "code": "Unknown",
        "category": "Pansharpening (Fusion)"
    },
    {
        "id": 3,
        "title": "S3Mamba-Pan: Spectral-Spatial-Scale Mamba With Frequency-Decoupled Dual-Stream for Pansharpening",
        "authors": "Song, Zishun; Zhang, Yao; Li, Haoyu; He, Yanlin; Zhao, Jiawei; Yang, Yi; Zhang, Wei; Wang, Dezhen",
        "journal": "IEEE Transactions on Geoscience and Remote Sensing (TGRS)",
        "year": 2026,
        "volume": 64,
        "pages": "5404816",
        "doi": "10.1109/TGRS.2026.3686021",
        "impact_factor": 9.4,
        "open_access": "No",
        "code": "Unknown",
        "category": "Pansharpening (Fusion)"
    },
    {
        "id": 4,
        "title": "Low-Rank Gradient Guidance With Mutual-Guided Mamba for Hyperspectral Pansharpening",
        "authors": "Wu, Chanyue; Wang, Dong; Mao, Hanyu; Bai, Zongwen; Li, Ying; Shang, Changjing; Shen, Qiang",
        "journal": "IEEE Journal of Selected Topics in Applied Earth Observations and Remote Sensing (JSTARS)",
        "year": 2026,
        "volume": 19,
        "pages": "5948-5966",
        "doi": "10.1109/jstars.2026.3655787",
        "impact_factor": 6.68,
        "open_access": "No",
        "code": "Unknown",
        "category": "Pansharpening (Fusion)"
    },
    {
        "id": 5,
        "title": "3SP-DUNet: An Interpretable Deep-Unfolded Network With Spatial-Spectral Sparse Priors for Hyperspectral Pansharpening",
        "authors": "Yang, Yong; Liu, Xuan; Huang, Shuying; Wang, Xiaozheng; Lu, Jinlin",
        "journal": "IEEE Transactions on Geoscience and Remote Sensing (TGRS)",
        "year": 2026,
        "volume": 64,
        "pages": "1-16",
        "doi": "10.1109/tgrs.2026.3674159",
        "impact_factor": 9.4,
        "open_access": "No",
        "code": "Unknown",
        "category": "Pansharpening (Fusion)"
    },
]

headers = [
    "No.", "Paper Title", "Authors", "Journal", "Year", "Volume", "Pages", 
    "DOI", "Impact Factor (JCR 2025)", "Open Access", "Code Available", "Category"
]

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for row_idx, paper in enumerate(papers, 2):
    row_data = [
        paper["id"],
        paper["title"],
        paper["authors"],
        paper["journal"],
        paper["year"],
        paper["volume"],
        paper["pages"],
        paper["doi"],
        paper["impact_factor"],
        paper["open_access"],
        paper["code"],
        paper["category"]
    ]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

col_widths = [5, 55, 45, 45, 8, 10, 15, 45, 22, 14, 30, 25]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:L{len(papers)+1}"

output_path = r"D:\academic\project1\IEEE_2026_HSI_Fusion_Papers.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print("Contains only the 5 papers directly on HSI pansharpening/fusion (classification papers removed).")