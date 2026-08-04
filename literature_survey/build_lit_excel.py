"""Build the 2026 IEEE HSI-fusion literature workbook.

Inclusion rules (as specified):
  * IEEE journals only
  * Published 2026 (2027 volumes do not exist yet as of Aug 2026)
  * Journal Impact Factor >= 7.0
  * Open Access = No (verified per-article via Unpaywall)
  * Topic: HSI-MSI fusion / hyperspectral super-resolution / pansharpening
"""
import json
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# JCR release of June 2026 (2025 citation data)
IF_TABLE = {
    "IEEE Transactions on Pattern Analysis and Machine Intelligence": (20.4, "Q1"),
    "IEEE Transactions on Image Processing": (15.3, "Q1"),
    "IEEE Transactions on Circuits and Systems for Video Technology": (10.8, "Q1"),
    "IEEE Transactions on Multimedia": (9.9, "Q1"),
    "IEEE Transactions on Neural Networks and Learning Systems": (9.7, "Q1"),
    "IEEE Transactions on Geoscience and Remote Sensing": (9.4, "Q1"),
}
ABBR = {
    "IEEE Transactions on Pattern Analysis and Machine Intelligence": "IEEE TPAMI",
    "IEEE Transactions on Image Processing": "IEEE TIP",
    "IEEE Transactions on Circuits and Systems for Video Technology": "IEEE TCSVT",
    "IEEE Transactions on Multimedia": "IEEE TMM",
    "IEEE Transactions on Neural Networks and Learning Systems": "IEEE TNNLS",
    "IEEE Transactions on Geoscience and Remote Sensing": "IEEE TGRS",
}

# doi-suffix -> (method family, core mechanism, priority, what we can borrow)
ANN = {
 "tpami.2026.3681688": ("Blind / unsupervised", "Stochastic learning of band-wise blur kernels; no paired training data", "HIGH", "Kernel-agnostic degradation estimation - directly attacks the CAVE->Harvard blur/SRF mismatch that broke Fusformer"),
 "tip.2026.3689415": ("Unfolding + generative prior", "Uncertainty-weighted generative prior inside a sparse model-guided solver", "HIGH", "Per-pixel uncertainty head; down-weight unreliable regions instead of hallucinating them under domain shift"),
 "tip.2026.3695389": ("Tensor unfolding (unsupervised)", "Self-expressive high-order tensor unrolling, trained without HR ground truth", "HIGH", "Unsupervised unrolling -> no CAVE-specific supervision to overfit; ideal OOD baseline"),
 "tip.2025.3647207": ("Diffusion (unsupervised)", "Coupled diffusion posterior sampling constrained by the observation model", "HIGH", "Posterior sampling keeps the physical forward model at inference - a robustness mechanism our 10 baselines lack"),
 "tip.2025.3636789": ("Unfolding + Transformer + wavelet", "Content-adaptive unfolding with wavelet-domain Transformer stages", "HIGH", "Content-adaptive step sizes per unfolding stage; wavelet split separates spectral low-freq from spatial high-freq"),
 "tip.2026.3657219": ("Equivariance", "Equivariant network for mosaiced + PAN fusion", "MEDIUM", "Equivariance constraints are a principled inductive bias for generalization; transferable to HSI-MSI"),
 "tgrs.2026.3703367": ("Blind / causal", "Causal degradation-guided network with spatial-frequency attention", "HIGH", "Causal decoupling of degradation from content - the exact failure mode of our Transformer baselines"),
 "tgrs.2026.3687252": ("MoE + Transformer + Mamba", "Degradation-aware mixture-of-experts routing over Transformer and Mamba experts", "HIGH", "Route by estimated degradation, so one model covers multiple domains without retraining"),
 "tgrs.2026.3680287": ("MoE", "Region-aware expert routing for HSI-MSI fusion", "HIGH", "Spatially adaptive capacity; Harvard's shadows/foliage need different experts than CAVE's flat textures"),
 "tgrs.2026.3718050": ("Unsupervised + registration", "Deformable bilinear fusion for unregistered HSI-MSI pairs", "HIGH", "Handles misalignment - a real-sensor failure our synthetic CAVE protocol never exposes"),
 "tgrs.2026.3694409": ("Mamba", "High-frequency dynamic guided Mamba, explicitly targeting robustness", "HIGH", "Linear-complexity global context without Transformer's data hunger - the core fix for Fusformer's collapse"),
 "tgrs.2026.3715225": ("Mamba", "Modal-synergy Mamba for cross-modal HSI-MSI interaction", "HIGH", "Cross-modal SSM scanning as a cheaper, better-regularized substitute for cross-attention"),
 "tgrs.2026.3699818": ("Mamba + tensor", "Block-term decomposition guiding frequency-domain Mamba modulation", "MEDIUM", "Low-rank tensor structure as an explicit constraint on the SSM feature space"),
 "tgrs.2025.3650662": ("Mamba + wavelet", "Wavelet-enhanced spatial-spectral prior injection into a Mamba backbone", "MEDIUM", "Prior injection points - where to inject physics into an SSM stack"),
 "tgrs.2026.3651526": ("Diffusion", "Spectral-degradation-guided diffusion Schrodinger bridge", "MEDIUM", "Bridge formulation starts from the LR-HSI rather than noise -> far fewer sampling steps"),
 "tgrs.2026.3664848": ("Diffusion (unsupervised)", "Hybrid-prior guided coupled diffusion, unsupervised", "MEDIUM", "Hybrid hand-crafted + learned priors; a hedge against pure data-driven overfitting"),
 "tgrs.2026.3656069": ("Diffusion + reference matching", "Mutual enhancement between reference matching and fusion", "MEDIUM", "Joint matching+fusion; relevant if the MSI is not perfectly co-registered"),
 "tgrs.2026.3674159": ("Deep unfolding", "Interpretable unfolded net with spatial-spectral sparse priors", "HIGH", "Sparse-prior unfolding - same family as DHIF-Net, our strongest OOD baseline"),
 "tgrs.2026.3703738": ("Unfolding + learnable prior", "Hierarchical progressive transform with a learnable prior", "MEDIUM", "Progressive coarse-to-fine unfolding stabilises very large scale factors (s=8..32)"),
 "tgrs.2026.3697926": ("Tensor decomposition", "Coupled tensor wheel decomposition", "MEDIUM", "Training-free spectral-fidelity anchor; useful as a physics regulariser"),
 "tgrs.2026.3689503": ("Tensor + regularization", "Joint low-rank and smooth tensor regularization", "MEDIUM", "Low-rank + smoothness penalty that can be added to any deep loss"),
 "tgrs.2026.3703141": ("Low-rank + Transformer", "Spectral-to-spatial aggregation Transformer over a low-rank representation", "MEDIUM", "Low-rank projection before attention cuts parameters and curbs overfitting"),
 "tgrs.2026.3691537": ("Frequency domain", "Progressive supervision with frequency-domain decoupling", "HIGH", "Deep supervision in the frequency domain - cheap, model-agnostic, improves spectral fidelity"),
 "tgrs.2026.3716019": ("Graph + frequency", "Graph-based spatial-frequency learning", "MEDIUM", "Graph relations generalise better than fixed grids across scene statistics"),
 "tgrs.2026.3708339": ("Progressive CNN-Transformer", "Local-global progressive fusion", "MEDIUM", "Explicit local/global split - keeps CNN locality bias while adding global context"),
 "tgrs.2026.3713929": ("Staged reconstruction", "Stagewise spectral-structural coordinated reconstruction", "MEDIUM", "Decouple spectral and structural stages so spatial sharpening cannot corrupt spectra"),
 "tgrs.2026.3659928": ("Attention", "Modal interaction with window dilation attention", "MEDIUM", "Dilated windows enlarge receptive field at low cost"),
 "tgrs.2026.3653545": ("Multistage CNN", "Spatial-spectral edge-enhancement multistage fusion", "LOW", "Edge-aware supervision for high-frequency recovery"),
 "tgrs.2026.3671284": ("Differential features", "S2 differential feature awareness network", "LOW", "Differential (residual-of-residual) features suppress redundancy"),
 "tgrs.2026.3683056": ("Detail injection", "Classical detail-injection framework for HSI/MSI/PAN", "MEDIUM", "Detail injection is inherently scale/domain robust - strong classical control arm"),
 "tgrs.2026.3672273": ("Random field ensemble", "Context-reinforced random-field ensembles at arbitrary resolution", "MEDIUM", "Arbitrary-resolution handling - our protocol is locked to s=8"),
 "tgrs.2026.3678075": ("Dual-branch CNN", "Dual-branch spatial/spectral network", "LOW", "Standard dual-stream control arm"),
 "tgrs.2026.3665830": ("Recurrent refinement", "Dynamic recurrent self-refinement", "MEDIUM", "Weight-shared recurrence: more effective depth at constant parameter count"),
 "tgrs.2026.3688486": ("Efficiency", "Spectrally-spatially coupled dynamic reduction", "MEDIUM", "Dynamic channel reduction - lets us report FLOPs/params, which the seniors' paper omits"),
 "tgrs.2026.3693935": ("Efficient Transformer", "Visual differential spatially-projected Transformer", "MEDIUM", "Spatial projection cuts attention cost for 512x512 Harvard scenes"),
 "tgrs.2026.3651549": ("GAN + KAN", "Low-rank spectral-spatial SR with a KAN-based GAN", "LOW", "KAN layers as a drop-in replacement for MLPs in the spectral branch"),
 "tcsvt.2025.3597548": ("Diffusion", "Hierarchically conditional diffusion for HSI reconstruction", "MEDIUM", "Hierarchical conditioning schedule - coarse spectra first, fine spatial later"),
 "tcsvt.2025.3621300": ("Self-supervised low-rank", "Self-supervised low-rank decomposition network", "HIGH", "Self-supervision means it can adapt on Harvard at test time with no labels"),
 "tcsvt.2025.3626779": ("RWKV", "Receptance weighted key-value (RWKV) linear-attention backbone", "MEDIUM", "Third linear-complexity family alongside Mamba/SSM - worth an ablation"),
 "tmm.2026.3668557": ("Topology / geometry", "Boundary perception and topology inference", "MEDIUM", "Topology constraints preserve object structure under heavy blur"),
 "tnnls.2026.3709066": ("Anisotropic CNN", "Anisotropic structure-aware, spectrally recalibrated network", "MEDIUM", "Spectral recalibration block - cheap SAM improvement, easy ablation win"),
 "tnnls.2025.3631243": ("Hybrid Transformer-CNN", "Butterfly residual net: spectral Transformer + depthwise convolutions", "LOW", "Depthwise convs restore locality bias inside a Transformer"),
}


def clean(t):
    t = re.sub(r"<[^>]+>", "", t)
    return re.sub(r"\s+", " ", t).strip()


cands = json.load(open("lit_candidates.json", encoding="utf-8"))
rows, excluded = [], []

for c in cands:
    jr = c["journal"]
    if jr not in IF_TABLE:
        continue
    key = c["doi"].split("/")[-1]
    if c.get("is_oa") is True:
        excluded.append((clean(c["title"]), ABBR[jr], c["doi"],
                         f"Open Access ({c.get('oa_status')})"))
        continue
    if c.get("is_oa") is None and key != "tip.2026.3714832":
        excluded.append((clean(c["title"]), ABBR[jr], c["doi"], "OA status unverified"))
        continue
    fam, mech, prio, borrow = ANN.get(
        key, ("Fusion network", "See paper", "MEDIUM", "To be assessed"))
    rows.append({
        "title": clean(c["title"]),
        "authors": c["authors"],
        "journal": jr,
        "abbr": ABBR[jr],
        "year": c["year"],
        "volume": c["volume"] or "Early Access",
        "pages": c["pages"] or "Early Access",
        "doi": c["doi"],
        "if": IF_TABLE[jr][0],
        "q": IF_TABLE[jr][1],
        "oa": "No",
        "oa_src": ("Unpaywall: closed" if c.get("is_oa") is False
                   else "Subscription journal; too recent for Unpaywall index"),
        "fam": fam, "mech": mech, "prio": prio, "borrow": borrow,
    })

order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
rows.sort(key=lambda r: (order[r["prio"]], -r["if"], r["title"]))

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Sheet 1
ws = wb.active
ws.title = "IEEE_2026_HSI_Fusion"

HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color="1F3864")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
BORDER = Border(*[Side(style="thin")] * 4)
PRIO_FILL = {
    "HIGH": PatternFill("solid", start_color="C6EFCE"),
    "MEDIUM": PatternFill("solid", start_color="FFEB9C"),
    "LOW": PatternFill("solid", start_color="F2F2F2"),
}

headers = ["No.", "Paper Title", "Authors", "Journal", "Journal (Abbr.)", "Year",
           "Volume", "Pages", "DOI", "Impact Factor (JCR 2026 rel.)", "Quartile",
           "Open Access", "OA Verification", "Method Family", "Core Mechanism",
           "Priority for Our Work", "What We Can Borrow", "Code Available"]

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, HDR_ALIGN, BORDER

for i, r in enumerate(rows, 1):
    data = [i, r["title"], r["authors"], r["journal"], r["abbr"], r["year"],
            r["volume"], r["pages"], r["doi"], r["if"], r["q"], r["oa"],
            r["oa_src"], r["fam"], r["mech"], r["prio"], r["borrow"], "Unknown"]
    for col, v in enumerate(data, 1):
        cell = ws.cell(i + 1, col, v)
        cell.alignment, cell.border = CELL_ALIGN, BORDER
        if col == 16:
            cell.fill = PRIO_FILL[r["prio"]]
            cell.font = Font(bold=True)

widths = [5, 62, 42, 46, 14, 7, 12, 14, 30, 15, 9, 12, 26, 24, 46, 12, 60, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:R{len(rows) + 1}"

# ---------------------------------------------------------------- Sheet 2
ws2 = wb.create_sheet("Selection_Criteria")
crit = [
    ["Criterion", "Setting", "How it was enforced"],
    ["Publisher", "IEEE only", "Crossref DOI prefix 10.1109 filter"],
    ["Publication year", "2026", "Crossref filter from-pub-date:2026-01-01. NOTE: as of Aug 2026 no "
     "2027-dated IEEE volumes exist; 2027 issues begin appearing Jan 2027."],
    ["Impact Factor", ">= 7.0", "JCR released June 2026 (2025 citation data). Journals below the "
     "threshold were dropped: JSTARS (~5.5-6.7), TCI (~4.2), GRSL (~4.0), IEEE Access (~3.4)."],
    ["Open Access", "No", "Per-article check via the Unpaywall API. Only articles returning "
     "is_oa = false were kept. Open-access hits are listed on the Excluded sheet."],
    ["Topic - Track 1 (sheet 1)", "HSI-MSI fusion / HS super-resolution / pansharpening",
     "Title screening. Classification, unmixing, detection and denoising papers removed."],
    ["Topic - Track 2 (sheet 4)", "HSI + LiDAR / SAR / multimodal joint classification fusion",
     "Separate Crossref harvest. Title must name a second modality (LiDAR, SAR, DSM, "
     "PAN, multimodal, cross-modal, multi-source) together with a spectral keyword."],
    ["", "", ""],
    ["Journal", "Impact Factor", "Quartile"],
    *[[j, v[0], v[1]] for j, v in sorted(IF_TABLE.items(), key=lambda x: -x[1][0])],
]
for ri, row in enumerate(crit, 1):
    for ci, v in enumerate(row, 1):
        cell = ws2.cell(ri, ci, v)
        cell.alignment = CELL_ALIGN
        if ri in (1, 9):
            cell.font, cell.fill = HDR_FONT, HDR_FILL
for i, w in enumerate([46, 20, 95], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------- Sheet 3
ws3 = wb.create_sheet("Excluded")
ws3.append(["Paper Title", "Journal", "DOI", "Reason for exclusion"])
for c in ws3[1]:
    c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
for e in excluded:
    ws3.append(list(e))
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.alignment = CELL_ALIGN
for i, w in enumerate([70, 16, 32, 34], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ------------------------------------------- Sheet 4: HSI + LiDAR / multimodal
OFFTOPIC = re.compile(r"land surface temperature|gps trajector", re.I)
FAM_RULES = [
    (r"mamba|state.?space", "Mamba / state-space",
     "Linear-complexity state-space scanning across modalities"),
    (r"masked self.?attention|self.?attention|transformer|former\b", "Transformer / attention",
     "Attention-based cross-modal token interaction"),
    (r"contrastive|positive matching|curriculum", "Contrastive learning",
     "Aligns modality embeddings before fusion instead of concatenating features"),
    (r"prompt|distillation|language|textual|dino", "Prompt / knowledge transfer",
     "Prompt tuning or distillation transfers knowledge across modalities"),
    (r"diffusion", "Diffusion", "Generative completion of missing modality features"),
    (r"graph", "Graph learning", "Graph-structured cross-modal relation modelling"),
    (r"uncertain|fuzzy|information bottleneck|dendritic", "Uncertainty / information theory",
     "Models modality reliability explicitly rather than trusting both equally"),
    (r"expert|moe", "Mixture-of-Experts", "Modality- or region-conditioned expert routing"),
    (r"contourlet|frequenc|amplitud|wavelet", "Frequency domain",
     "Frequency-decomposed fusion of spatial and elevation cues"),
]
HIGH_RE = re.compile(r"masked self.?attention|mamba|uncertain|information bottleneck|"
                     r"contrastive|expert|frequenc", re.I)

mm = json.load(open("lit_multimodal.json", encoding="utf-8"))
ws4 = wb.create_sheet("IEEE_2026_HSI_LiDAR_Multimodal")
h4 = ["No.", "Paper Title", "Authors", "Journal", "Journal (Abbr.)", "Year", "Volume",
      "Pages", "DOI", "Impact Factor (JCR 2026 rel.)", "Quartile", "Open Access",
      "OA Verification", "Method Family", "Core Mechanism", "Priority for Our Work"]
for col, h in enumerate(h4, 1):
    c = ws4.cell(1, col, h)
    c.font, c.fill, c.alignment, c.border = HDR_FONT, HDR_FILL, HDR_ALIGN, BORDER

n = 0
mm_rows = []
for m in mm:
    if m.get("is_oa") is not False or OFFTOPIC.search(m["title"]):
        if m.get("is_oa") is True:
            excluded.append((clean(m["title"]), ABBR.get(m["journal"], m["journal"]),
                             m["doi"], f"Open Access ({m.get('oa_status')})"))
        continue
    t = clean(m["title"])
    fam, mech = "Multimodal fusion network", "Cross-modal feature fusion for joint classification"
    for pat, f, mo in FAM_RULES:
        if re.search(pat, t, re.I):
            fam, mech = f, mo
            break
    mm_rows.append((t, m, fam, mech, "HIGH" if HIGH_RE.search(t) else "MEDIUM"))

mm_rows.sort(key=lambda x: (order[x[4]], -IF_TABLE[x[1]["journal"]][0], x[0]))
for n, (t, m, fam, mech, prio) in enumerate(mm_rows, 1):
    jr = m["journal"]
    data = [n, t, m["authors"], jr, ABBR[jr], m["year"], m["volume"] or "Early Access",
            m["pages"] or "Early Access", m["doi"], IF_TABLE[jr][0], IF_TABLE[jr][1],
            "No", "Unpaywall: closed", fam, mech, prio]
    for col, v in enumerate(data, 1):
        c = ws4.cell(n + 1, col, v)
        c.alignment, c.border = CELL_ALIGN, BORDER
        if col == 16:
            c.fill, c.font = PRIO_FILL[prio], Font(bold=True)
for i, w in enumerate([5, 72, 42, 46, 14, 7, 12, 14, 30, 15, 9, 12, 20, 26, 52, 12], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "B2"
ws4.auto_filter.ref = f"A1:P{n + 1}"

# rebuild Excluded sheet now that multimodal exclusions are known
wb.remove(ws3)
ws3 = wb.create_sheet("Excluded")
ws3.append(["Paper Title", "Journal", "DOI", "Reason for exclusion"])
for c in ws3[1]:
    c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
for e in excluded:
    ws3.append(list(e))
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.alignment = CELL_ALIGN
for i, w in enumerate([70, 16, 32, 34], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

wb.save("IEEE_2026_HSI_Fusion_Literature.xlsx")
print("multimodal rows:", n)
print(f"Saved {len(rows)} papers; {len(excluded)} excluded.")
import collections
print(collections.Counter(r["abbr"] for r in rows))
print(collections.Counter(r["prio"] for r in rows))
